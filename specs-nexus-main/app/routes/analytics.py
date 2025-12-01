import logging
from datetime import datetime, timedelta
from typing import Dict, Any, Optional
from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy import func, or_, and_, text
from sqlalchemy.orm import Session
from sqlalchemy.exc import ProgrammingError, OperationalError
from pydantic import BaseModel
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.database import SessionLocal
from app import models, schemas
from app.auth_utils import get_current_officer

logger = logging.getLogger("app.analytics")

router = APIRouter(prefix="/analytics", tags=["Analytics"])

def get_db():
    db = SessionLocal()
    try:
        # Test database connection
        db.execute(text("SELECT 1"))
        logger.debug("Database connection successful")
        # Check if users table exists (MySQL-compatible)
        result = db.execute(
            text("""
                SELECT COUNT(*) 
                FROM information_schema.tables 
                WHERE table_name = 'users'
            """)
        ).scalar()
        if not result:
            logger.error("Users table does not exist in the database")
            raise HTTPException(status_code=500, detail="Users table not found in the database")
        logger.debug("Users table exists")
        yield db
    except (ProgrammingError, OperationalError) as e:
        logger.error(f"Database error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        logger.error(f"Unexpected error in get_db: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")
    finally:
        db.close()

class DateRangeFilter(BaseModel):
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    include_archived: Optional[bool] = False  # New parameter to include archived records

@router.get("/dashboard", response_model=dict)
def get_dashboard_data(date_filter: Optional[DateRangeFilter] = None, db: Session = Depends(get_db)) -> Dict[str, Any]:
    logger.debug("Starting dashboard data aggregation")
    
    try:
        # Apply date range filter if provided
        date_filter = date_filter or DateRangeFilter()
        start_date = date_filter.start_date or (datetime.now() - timedelta(days=365 * 2))  # Extended to 2 years
        end_date = date_filter.end_date or datetime.now()
        include_archived = date_filter.include_archived or False

        # Validate date range
        if start_date > end_date:
            raise HTTPException(status_code=400, detail="Start date must be before end date")

        # Total BSCS Students (all users in the system)
        try:
            total_cs_students = db.query(models.User).count()
            logger.debug(f"Total BSCS students: {total_cs_students}")
        except ProgrammingError as e:
            logger.error(f"Error querying users table: {str(e)}")
            total_cs_students = 0
            raise HTTPException(status_code=500, detail=f"Error querying users table: {str(e)}")

        no_users = total_cs_students == 0
        if no_users:
            logger.warning("No users found in the users table")

        # Total Specs Members (distinct users with payment_status="Paid" within the date range)
        total_specs_members_query = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        )
        total_specs_members = total_specs_members_query.distinct().count()
        logger.debug(f"Total Specs members: {total_specs_members}")

        # Semester-specific Specs members
        total_specs_members_first_sem = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).distinct().count()
        total_specs_members_second_sem = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).distinct().count()
        logger.debug(f"1st Semester Specs members: {total_specs_members_first_sem}, 2nd Semester Specs members: {total_specs_members_second_sem}")

        # None Specs members (users with payment_status="Not Paid" or "Verifying" within the date range)
        none_specs = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status.in_(["Not Paid", "Verifying"]),
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).distinct().count()
        none_specs_first_sem = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status.in_(["Not Paid", "Verifying"]),
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).distinct().count()
        none_specs_second_sem = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status.in_(["Not Paid", "Verifying"]),
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).distinct().count()
        logger.debug(f"None Specs: {none_specs}, 1st Sem: {none_specs_first_sem}, 2nd Sem: {none_specs_second_sem}")

        # Members by requirement (for charts)
        members_by_requirement_raw = db.query(
            models.Clearance.requirement,
            func.count(func.distinct(models.Clearance.user_id))
        ).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).group_by(models.Clearance.requirement).all()
        members_by_requirement = {req: count for req, count in members_by_requirement_raw}
        logger.debug(f"Members by requirement: {members_by_requirement}")

        # Active members (last 30 days) and recent activity (last 7 days)
        thirty_days_ago = datetime.now() - timedelta(days=30)
        seven_days_ago = datetime.now() - timedelta(days=7)
        
        try:
            active_members = db.query(models.User).filter(
                models.User.last_active >= thirty_days_ago,
                models.User.last_active.isnot(None)
            ).count()
            
            recent_activity = db.query(models.User).filter(
                models.User.last_active >= seven_days_ago,
                models.User.last_active.isnot(None)
            ).count()
            
            inactive_members = total_cs_students - active_members
            logger.debug(f"Active members: {active_members}, Inactive members: {inactive_members}, Recent activity (7 days): {recent_activity}")
        except ProgrammingError as e:
            logger.error(f"Error querying active/inactive members: {str(e)}")
            active_members = 0
            inactive_members = total_cs_students
            recent_activity = 0

        # Payment status counts (overall and semester-specific)
        not_paid_count = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Not Paid",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        verifying_count = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Verifying",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        paid_count = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).count()

        not_paid_first_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Not Paid",
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        not_paid_second_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Not Paid",
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        verifying_first_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Verifying",
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        verifying_second_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Verifying",
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        paid_first_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).count()
        paid_second_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).count()
        logger.debug(f"Payment Analytics - Not Paid: {not_paid_count}, Verifying: {verifying_count}, Paid: {paid_count}")
        logger.debug(f"1st Sem - Not Paid: {not_paid_first_sem}, Verifying: {verifying_first_sem}, Paid: {paid_first_sem}")
        logger.debug(f"2nd Sem - Not Paid: {not_paid_second_sem}, Verifying: {verifying_second_sem}, Paid: {paid_second_sem}")

        # Payment methods and trends
        payment_methods = db.query(
            models.Clearance.payment_method,
            func.count(models.Clearance.id)
        ).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_method.isnot(None),
            or_(
                and_(
                    models.Clearance.payment_status == "Paid",
                    models.Clearance.payment_date >= start_date,
                    models.Clearance.payment_date <= end_date
                ),
                models.Clearance.payment_status.in_(["Not Paid", "Verifying"])
            )
        ).group_by(models.Clearance.payment_method).all()
        logger.debug(f"Raw payment methods query result: {payment_methods}")
        preferred_payment_methods = [{"method": method, "count": count, "firstSemCount": 0, "secondSemCount": 0} for method, count in payment_methods]

        payment_method_trends = db.query(
            models.Clearance.payment_method,
            models.Clearance.requirement,
            func.count(models.Clearance.id).label('count')
        ).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_method.isnot(None),
            or_(
                and_(
                    models.Clearance.payment_status == "Paid",
                    models.Clearance.payment_date >= start_date,
                    models.Clearance.payment_date <= end_date
                ),
                models.Clearance.payment_status.in_(["Not Paid", "Verifying"])
            )
        ).group_by(models.Clearance.payment_method, models.Clearance.requirement).all()
        logger.debug(f"Raw payment method trends query result: {payment_method_trends}")
        payment_method_trends_dict = {}
        for method, requirement, count in payment_method_trends:
            if method not in payment_method_trends_dict:
                payment_method_trends_dict[method] = {"firstSemCount": 0, "secondSemCount": 0}
            if requirement == "1st Semester Membership":
                payment_method_trends_dict[method]["firstSemCount"] = count
            elif requirement == "2nd Semester Membership":
                payment_method_trends_dict[method]["secondSemCount"] = count
        payment_method_trends_list = [
            {"method": method, "firstSemCount": data["firstSemCount"], "secondSemCount": data["secondSemCount"]}
            for method, data in payment_method_trends_dict.items()
        ]
        for method in preferred_payment_methods:
            for trend in payment_method_trends_list:
                if trend["method"] == method["method"]:
                    method["firstSemCount"] = trend["firstSemCount"]
                    method["secondSemCount"] = trend["secondSemCount"]
        logger.debug(f"Preferred payment methods: {preferred_payment_methods}")
        logger.debug(f"Payment method trends: {payment_method_trends_list}")

        # Payment details by requirement and year
        payment_by_req_year_raw = db.query(
            models.User.year,
            models.Clearance.requirement,
            models.Clearance.payment_status,
            func.count(models.Clearance.id)
        ).join(models.Clearance, models.Clearance.user_id == models.User.id).filter(
            models.Clearance.archived == False,
            or_(
                and_(
                    models.Clearance.payment_status == "Paid",
                    models.Clearance.payment_date >= start_date,
                    models.Clearance.payment_date <= end_date
                ),
                and_(
                    models.Clearance.payment_status.in_(["Not Paid", "Verifying"]),
                    models.Clearance.last_updated >= start_date,
                    models.Clearance.last_updated <= end_date
                )
            )
        ).group_by(models.User.year, models.Clearance.requirement, models.Clearance.payment_status).all()

        byRequirementAndYear = {}
        for user_year, requirement, payment_status, count in payment_by_req_year_raw:
            if not user_year:
                user_year = "Unspecified"
            if requirement not in byRequirementAndYear:
                byRequirementAndYear[requirement] = {}
            if user_year not in byRequirementAndYear[requirement]:
                byRequirementAndYear[requirement][user_year] = {"Not Paid": 0, "Verifying": 0, "Paid": 0}
            byRequirementAndYear[requirement][user_year][payment_status] = count
        logger.debug(f"Payment details by requirement and year: {byRequirementAndYear}")

        # Event details and participation rates
        events_query = db.query(models.Event).filter(
            models.Event.archived == include_archived,  # Use include_archived parameter
            or_(
                models.Event.date >= start_date,
                models.Event.date <= end_date,
                models.Event.date.is_(None)  # Include NULL dates
            )
        )
        events = events_query.all()
        logger.debug(f"Raw events query result: {[(e.title, e.date, e.participant_count, e.archived) for e in events]}")
        events_engagement = []
        events_by_year = {}
        for event in events:
            event_year = event.date.year if event.date else "Unknown"
            engagement = {
                "title": event.title,
                "participant_count": event.participant_count or 0,
                "participation_rate": round((event.participant_count / total_specs_members) * 100, 2) if total_specs_members > 0 and event.participant_count else 0
            }
            events_engagement.append(engagement)
            if event_year not in events_by_year:
                events_by_year[event_year] = []
            events_by_year[event_year].append(engagement)
        popular_events = sorted(events_engagement, key=lambda x: x["participant_count"], reverse=True)[:5]
        logger.debug(f"Event engagement: {events_engagement}")
        logger.debug(f"Popular events: {popular_events}")

        # Clearance by requirement
        clearance_by_requirement_raw = db.query(
            models.Clearance.requirement,
            models.Clearance.status,
            func.count(models.Clearance.id)
        ).filter(
            models.Clearance.archived == False,
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).group_by(models.Clearance.requirement, models.Clearance.status).all()
        clearance_tracking = {}
        for requirement, status, count in clearance_by_requirement_raw:
            if requirement not in clearance_tracking:
                clearance_tracking[requirement] = {"Clear": 0, "Processing": 0, "Not Yet Cleared": 0}
            clearance_tracking[requirement][status] = count
        logger.debug(f"Clearance tracking: {clearance_tracking}")

        # Compliance by year
        compliance_by_year = db.query(
            models.User.year,
            models.Clearance.status,
            func.count(models.Clearance.id)
        ).join(models.Clearance, models.Clearance.user_id == models.User.id).filter(
            models.Clearance.archived == False,
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).group_by(models.User.year, models.Clearance.status).all()

        compliance = {}
        for year, status, count in compliance_by_year:
            year = year or "Unspecified"
            if year not in compliance:
                compliance[year] = {"Clear": 0, "Processing": 0, "Not Yet Cleared": 0}
            compliance[year][status] = count
        logger.debug(f"Compliance by year: {compliance}")

        logger.info("Dashboard data aggregated successfully")
        return {
            "membershipInsights": {
                "totalBSCSStudents": total_cs_students,
                "totalSpecsMembers": total_specs_members,
                "totalSpecsMembersFirstSem": total_specs_members_first_sem,
                "totalSpecsMembersSecondSem": total_specs_members_second_sem,
                "noneSpecs": none_specs,
                "noneSpecsFirstSem": none_specs_first_sem,
                "noneSpecsSecondSem": none_specs_second_sem,
                "activeMembers": active_members,
                "inactiveMembers": inactive_members,
                "recentActivityLast7Days": recent_activity,
                "membersByRequirement": members_by_requirement,
                "noUsers": no_users
            },
            "paymentAnalytics": {
                "byRequirementAndYear": byRequirementAndYear,
                "notPaid": not_paid_count,
                "verifying": verifying_count,
                "paid": paid_count,
                "notPaidFirstSem": not_paid_first_sem,
                "notPaidSecondSem": not_paid_second_sem,
                "verifyingFirstSem": verifying_first_sem,
                "verifyingSecondSem": verifying_second_sem,
                "paidFirstSem": paid_first_sem,
                "paidSecondSem": paid_second_sem,
                "preferredPaymentMethods": preferred_payment_methods,
                "paymentMethodTrends": payment_method_trends_list
            },
            "eventsEngagement": {
                "events": events_engagement,
                "popularEvents": popular_events,
                "breakdownByYear": events_by_year
            },
            "clearanceTracking": {
                "byRequirement": clearance_tracking,
                "complianceByYear": compliance
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error aggregating dashboard data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error while aggregating dashboard data: {str(e)}")

@router.get("/report/officer-dashboard")
def generate_officer_dashboard_report(
    db: Session = Depends(get_db),
    current_officer: models.Officer = Depends(get_current_officer)
):
    """Generate Excel report for officer dashboard"""
    logger.debug(f"Officer {current_officer.id} generating dashboard report")
    
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Get dashboard data by calling the endpoint logic directly
        start_date = datetime.now() - timedelta(days=365 * 2)
        end_date = datetime.now()
        include_archived = False
        
        # Reuse the same logic from get_dashboard_data
        total_cs_students = db.query(models.User).count()
        total_specs_members = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).distinct().count()
        
        thirty_days_ago = datetime.now() - timedelta(days=30)
        active_members = db.query(models.User).filter(
            models.User.last_active >= thirty_days_ago,
            models.User.last_active.isnot(None)
        ).count()
        inactive_members = total_cs_students - active_members
        
        payment_analytics_data = db.query(
            models.Clearance.payment_method,
            func.count(models.Clearance.id)
        ).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_method.isnot(None),
            or_(
                and_(
                    models.Clearance.payment_status == "Paid",
                    models.Clearance.payment_date >= start_date,
                    models.Clearance.payment_date <= end_date
                ),
                models.Clearance.payment_status.in_(["Not Paid", "Verifying"])
            )
        ).group_by(models.Clearance.payment_method).all()
        
        events_data = db.query(models.Event).filter(
            models.Event.archived == include_archived
        ).all()
        
        clearance_data = db.query(
            models.Clearance.requirement,
            models.Clearance.status,
            func.count(models.Clearance.id)
        ).filter(
            models.Clearance.archived == False,
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).group_by(models.Clearance.requirement, models.Clearance.status).all()
        
        # Sheet 1: Student Insights
        ws_students = wb.create_sheet("Student Insights")
        ws_students.append(["Student Insights Report"])
        ws_students.append([])
        ws_students.append(["Total BSCS Students", total_cs_students])
        ws_students.append(["Active Members", active_members])
        ws_students.append(["Inactive Members", inactive_members])
        ws_students.append([])
        
        # Get all users
        users = db.query(models.User).all()
        ws_students.append(["ID", "Name", "Student Number", "Year", "Block", "Email", "Status"])
        for user in users:
            is_active = user.last_active and user.last_active >= thirty_days_ago
            ws_students.append([
                user.id,
                user.full_name or "N/A",
                user.student_number or "N/A",
                user.year or "N/A",
                user.block or "N/A",
                user.email or "N/A",
                "Active" if is_active else "Inactive"
            ])
        
        # Sheet 2: Payment Analytics
        ws_payments = wb.create_sheet("Payment Analytics")
        ws_payments.append(["Payment Analytics Report"])
        ws_payments.append([])
        
        # Calculate payment stats
        verifying_count = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Verifying",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        paid_count = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).count()
        not_paid_count = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Not Paid",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        
        ws_payments.append(["Verifying", verifying_count])
        ws_payments.append(["Paid", paid_count])
        ws_payments.append(["Not Paid", not_paid_count])
        ws_payments.append([])
        ws_payments.append(["Payment Method", "Count"])
        for method, count in payment_analytics_data:
            ws_payments.append([method or "N/A", count])
        
        # Sheet 3: Events
        ws_events = wb.create_sheet("Events")
        ws_events.append(["Events Engagement Report"])
        ws_events.append([])
        ws_events.append(["Event Title", "Date", "Location", "Participant Count"])
        for event in events_data:
            ws_events.append([
                event.title or "N/A",
                event.date.strftime("%Y-%m-%d %H:%M") if event.date else "N/A",
                event.location or "N/A",
                event.participant_count or 0
            ])
        
        # Sheet 4: Clearance Tracking
        ws_clearance = wb.create_sheet("Clearance Tracking")
        ws_clearance.append(["Clearance Tracking Report"])
        ws_clearance.append([])
        ws_clearance.append(["Requirement", "Status", "Count"])
        for requirement, status, count in clearance_data:
            ws_clearance.append([
                requirement or "N/A",
                status or "N/A",
                count
            ])
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Officer {current_officer.id} generated dashboard report successfully")
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=officer-dashboard-report-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            }
        )
        
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")